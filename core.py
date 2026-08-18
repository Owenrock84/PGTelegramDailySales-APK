"""Shared PostgreSQL -> report files -> Telegram logic."""
from __future__ import annotations

import csv
import datetime as dt
import json
import re
import ssl
import tempfile
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import pg8000.dbapi
import requests
from openpyxl import Workbook, load_workbook


WRITE_WORDS = re.compile(r"\b(insert|update|delete|drop|alter|truncate|create|grant|revoke|copy|call|do|merge|vacuum|reindex)\b", re.I)


def validate_read_only_sql(sql: str) -> str:
    query = sql.strip()
    if query.endswith(";"):
        query = query[:-1].rstrip()
    if not query or not re.match(r"^(select|with)\b", query, re.I):
        raise ValueError("Only SELECT or read-only WITH queries are allowed.")
    if ";" in query:
        raise ValueError("Only one SQL statement is allowed per query box.")
    if WRITE_WORDS.search(query):
        raise ValueError("A write/administration SQL keyword was detected. Use read-only SELECT queries.")
    return query


def run_queries(config: dict):
    db = config["database"]
    ssl_mode = db.get("ssl", "require")
    ssl_context = None
    if ssl_mode == "require":
        ssl_context = ssl.create_default_context()
    elif ssl_mode == "allow-self-signed":
        ssl_context = ssl._create_unverified_context()
    conn = pg8000.dbapi.connect(
        host=db["host"], port=int(db.get("port", 5432)), database=db["name"],
        user=db["user"], password=db["password"], ssl_context=ssl_context,
        timeout=30,
    )
    try:
        conn.autocommit = False
        cur = conn.cursor()
        cur.execute("SET TRANSACTION READ ONLY")
        results = []
        for number in (1, 2):
            query = validate_read_only_sql(config[f"query{number}"])
            cur.execute(query)
            columns = [item[0] for item in cur.description]
            rows = cur.fetchall()
            results.append((f"Query {number}", columns, rows))
        conn.rollback()
        return results
    finally:
        conn.close()


def create_reports(results, output_format="xlsx", folder=None):
    report_dir = Path(folder or tempfile.mkdtemp(prefix="pgtelegram_"))
    report_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y-%m-%d_%H%M")
    files = []
    if output_format in ("xlsx", "both"):
        path = report_dir / f"PostgreSQL_Report_{stamp}.xlsx"
        wb = Workbook(); wb.remove(wb.active)
        for title, columns, rows in results:
            ws = wb.create_sheet(title[:31]); ws.append(columns)
            for row in rows: ws.append([safe_cell(value) for value in row])
            ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        wb.save(path); files.append(path)
    if output_format in ("csv", "both"):
        for index, (_title, columns, rows) in enumerate(results, 1):
            path = report_dir / f"Query_{index}_{stamp}.csv"
            with path.open("w", newline="", encoding="utf-8-sig") as stream:
                writer = csv.writer(stream); writer.writerow(columns); writer.writerows(rows)
            files.append(path)
    return files


def resource_path(filename: str) -> Path:
    """Resolve a bundled file in source, PyInstaller, or Android builds."""
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / filename


def _single_sales_row(result, label):
    _title, _columns, rows = result
    if len(rows) != 1:
        raise ValueError(f"{label} must return exactly one row; received {len(rows)} rows.")
    if len(rows[0]) != 5:
        raise ValueError(f"{label} must return exactly five columns: Bet, Win, Hand, Gross, Profit.")
    return list(rows[0])


def fill_daily_sales_template(results, folder=None, report_date=None):
    """Fill Yesterday B3:F3 and MTD B4:F4 without changing template styling."""
    if len(results) != 2:
        raise ValueError("Exactly two query results are required.")
    yesterday = _single_sales_row(results[0], "Query 1 (Yesterday)")
    mtd = _single_sales_row(results[1], "Query 2 (MTD)")
    report_date = report_date or (dt.date.today() - dt.timedelta(days=1))
    out_dir = Path(folder or tempfile.mkdtemp(prefix="pgtelegram_")); out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"Daily_Sales_{report_date:%Y-%m-%d}.xlsx"
    wb = load_workbook(resource_path("Daily Sales Template.xlsx"))
    ws = wb["Sheet1"]
    for col, value in enumerate(yesterday, 2): ws.cell(3, col).value = value
    for col, value in enumerate(mtd, 2): ws.cell(4, col).value = value
    ws["I3"] = dt.datetime.combine(report_date, dt.time())
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(path)
    return path, build_daily_sales_message(yesterday, mtd, report_date)


def _whole(value):
    if value is None: return "0"
    number = Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return f"{number:,.0f}"


def build_daily_sales_message(yesterday, mtd, report_date=None):
    report_date = report_date or (dt.date.today() - dt.timedelta(days=1))
    bet, win, hand, gross, profit = yesterday
    mtd_bet, mtd_win, mtd_hand, mtd_gross, mtd_profit = mtd
    return "\n".join([
        report_date.strftime("%A, %B %d, %Y").replace(" 0", " "),
        "",
        f"Hand : {_whole(hand)} Times✅",
        f"Stakes : MYR {_whole(bet)} ✅",
        f"Win Amount : MYR {_whole(win)} ✅",
        f"Company W/L : MYR {_whole(gross)} ✅",
        f"GGR (4%) : MYR {_whole(profit)} ✅",
        "",
        f"MTD Hand : {_whole(mtd_hand)} Times✅",
        f"MTD Stakes : MYR {_whole(mtd_bet)} ✅",
        f"MTD Win Amount : MYR {_whole(mtd_win)} ✅",
        f"MTD Company W/L : MYR {_whole(mtd_gross)} ✅",
        f"MTD GGR (4%) : MYR {_whole(mtd_profit)} ✅",
    ])


def safe_cell(value):
    if isinstance(value, (dict, list, tuple)): return json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, (bytes, bytearray)): return value.hex()
    if isinstance(value, dt.datetime) and value.tzinfo: return value.replace(tzinfo=None)
    return value


def send_bot_files(token: str, targets: list[str], files: list[Path], caption: str, log=print):
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    for target in targets:
        chat_id, topic_id = parse_target_spec(target)
        for path in files:
            with path.open("rb") as stream:
                data = {"chat_id": chat_id, "caption": caption}
                if topic_id is not None: data["message_thread_id"] = topic_id
                response = requests.post(url, data=data, files={"document": (path.name, stream)}, timeout=60)
            if not response.ok:
                raise RuntimeError(f"Telegram Bot API error for {target}: {response.text[:300]}")
        log(f"Sent report to {target}")


def send_bot_message(token: str, targets: list[str], message: str, log=print):
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    for target in targets:
        chat_id, topic_id = parse_target_spec(target)
        data = {"chat_id": chat_id, "text": message}
        if topic_id is not None: data["message_thread_id"] = topic_id
        response = requests.post(url, data=data, timeout=60)
        if not response.ok:
            raise RuntimeError(f"Telegram Bot API error for {target}: {response.text[:300]}")
        log(f"Sent Daily Sales message to {target}")


def parse_targets(text: str):
    return [item.strip() for item in re.split(r"[,\n]", text) if item.strip()]


def parse_target_spec(target: str):
    """Parse @group or @group|topic_id for a Telegram forum topic."""
    value = target.strip()
    if "|" not in value:
        return _coerce_peer(value), None
    chat_id, topic_text = value.rsplit("|", 1)
    if not chat_id.strip() or not topic_text.strip().isdigit():
        raise ValueError(f"Invalid topic target '{target}'. Expected @group|topic_id.")
    return _coerce_peer(chat_id.strip()), int(topic_text.strip())


def _coerce_peer(value: str):
    """Telethon requires numeric peer IDs as integers, not numeric strings."""
    text = value.strip()
    return int(text) if re.fullmatch(r"-?\d+", text) else text


def next_run_time(hour_minute: str, now=None):
    now = now or dt.datetime.now().astimezone()
    hour, minute = [int(item) for item in hour_minute.split(":", 1)]
    candidate = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if candidate <= now: candidate += dt.timedelta(days=1)
    return candidate
